#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import openpyxl as xl

print("[make_JSON.py]")

workbook = xl.load_workbook('Family_Tree_Responses.xlsx')
sheet = workbook['Form Responses 1']

membas = []
memCount = 0

X = sheet.max_row
Y = sheet.max_column

attributes = ["name", "father_name", "mother_name", "spouse_name", "s"]
attributeCount = 0
# print(len(attributes))
idCount = 1

for rows in range(2, X+1):
    attributeCount = 0
    membas.append({})
    memba = membas[memCount]
    memba["key"] = idCount
    idCount += 1
    for cols in range(2, Y + 1):
        attribute = attributes[attributeCount]
        data = str(sheet.cell(row=rows, column=cols).value)
        memba[attribute] = data
        attributeCount += 1
    memCount += 1

json.dump(membas, open("test.json", "w"))

importedFile = json.load(open("test.json"))
